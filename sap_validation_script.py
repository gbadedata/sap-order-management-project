"""
=============================================================================
SAP Order Management Project — Python Validation & Discrepancy Detection Script
Author  : Oluwagbade Joseph Odimayo
Project : SAP SD Order Management & Data Integrity Initiative
Period  : January – March 2025
Version : 1.0
=============================================================================

PURPOSE
-------
This script replicates the data validation logic applied in SAP SD order
processing, specifically the cross-checks that an order processor performs
using transaction codes VA03 (order review) and VK13 (pricing condition
records). It automates error detection, root cause classification, and
produces a structured discrepancy report — demonstrating the process
improvement work completed during the project.

HOW IT MAPS TO SAP WORKFLOW
----------------------------
  VA01 → Orders loaded from Order Register (order creation equivalent)
  VK13 → Pricing Master sheet (condition record baseline)
  VA03 → Script reads and validates each order (review equivalent)
  VA02 → Discrepancies flagged for manual correction (change order equivalent)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import os

# ── CONFIGURATION ─────────────────────────────────────────────────────────
INPUT_FILE    = "SAP_Order_Management_Project_Joseph_Odimayo.xlsx"
OUTPUT_REPORT = "SAP_Validation_Report_Joseph_Odimayo.xlsx"
PRICE_TOLERANCE_PCT = 0.05   # Flag if entered price deviates >5% from standard
DATE_FORMAT         = "%d/%m/%Y"

# Colour palette (consistent with main project workbook)
NAVY       = "1B3A5C"
BLUE       = "2E75B6"
WHITE      = "FFFFFF"
GREEN      = "1E7145"
GREEN_FILL = "C6EFCE"
RED        = "9C0006"
RED_FILL   = "FFC7CE"
AMBER      = "9C5700"
AMBER_FILL = "FFEB9C"
LIGHT_GREY = "F2F2F2"
ALT_ROW    = "EBF3FB"
DARK_GREY  = "2C2C2C"

# ── HELPERS ───────────────────────────────────────────────────────────────
def thin(color="BDD7EE"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=10, color=WHITE):
    return Font(name="Arial", bold=True, size=size, color=color)

def bfont(size=10, bold=False, color=DARK_GREY):
    return Font(name="Arial", size=size, bold=bold, color=color)

def pfill(color):
    return PatternFill("solid", fgColor=color)

def centre():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def right_al():
    return Alignment(horizontal="right", vertical="center")


# ── STEP 1: LOAD DATA ─────────────────────────────────────────────────────
print("=" * 65)
print("SAP ORDER VALIDATION SCRIPT")
print(f"Run Date : {datetime.now().strftime('%d %B %Y  %H:%M')}")
print(f"Input    : {INPUT_FILE}")
print("=" * 65)

print("\n[1/5] Loading data from workbook...")

try:
    orders_df = pd.read_excel(INPUT_FILE, sheet_name="Order Register",
                               header=1, engine="openpyxl")
    pricing_df = pd.read_excel(INPUT_FILE, sheet_name="Pricing Master (VK13)",
                                header=1, engine="openpyxl")
except FileNotFoundError:
    print(f"\n  ERROR: Cannot find '{INPUT_FILE}'.")
    print("  Ensure the main project workbook is in the same directory.")
    sys.exit(1)

# Clean column names
orders_df.columns  = [str(c).strip() for c in orders_df.columns]
pricing_df.columns = [str(c).strip() for c in pricing_df.columns]

# Drop any totals/summary rows (non-order rows at bottom)
orders_df = orders_df[orders_df["Sales Order No."].astype(str).str.startswith("SO-")].copy()
orders_df.reset_index(drop=True, inplace=True)

print(f"  ✓ {len(orders_df)} orders loaded from Order Register")
print(f"  ✓ {len(pricing_df)} pricing records loaded from Pricing Master (VK13)")


# ── STEP 2: BUILD PRICING LOOKUP (VK13 equivalent) ────────────────────────
print("\n[2/5] Building pricing condition lookup (VK13)...")

pricing_lookup = dict(
    zip(pricing_df["Material Code"].astype(str).str.strip(),
        pricing_df["Standard Price (£)"].astype(float))
)
print(f"  ✓ {len(pricing_lookup)} condition records indexed")


# ── STEP 3: RUN VALIDATION CHECKS ─────────────────────────────────────────
print("\n[3/5] Running validation checks across all orders...")

results = []
errors_found = 0

for idx, row in orders_df.iterrows():
    order_no     = str(row.get("Sales Order No.", "")).strip()
    po_ref       = str(row.get("PO Reference", "")).strip()
    order_date   = str(row.get("Order Date", "")).strip()
    delivery_date= str(row.get("Delivery Date", "")).strip()
    cust_id      = str(row.get("Customer ID", "")).strip()
    cust_name    = str(row.get("Customer Name", "")).strip()
    mat_code     = str(row.get("Material Code", "")).strip()
    description  = str(row.get("Product Description", "")).strip()
    uom          = str(row.get("UoM", "")).strip()
    week         = str(row.get("Week", "")).strip()

    try:
        qty           = float(row.get("Quantity", 0) or 0)
        entered_price = float(row.get("Entered Price (£)", 0) or 0)
        line_value    = float(row.get("Line Value (£)", 0) or 0)
    except (ValueError, TypeError):
        qty = entered_price = line_value = 0.0

    # --- Check 1: Price vs VK13 condition record ---
    standard_price = pricing_lookup.get(mat_code)
    price_flag     = False
    price_deviation = 0.0
    price_note     = "OK – matches VK13 condition record"

    if standard_price is None:
        price_flag  = True
        price_note  = "WARN – Material code not found in Pricing Master (VK13)"
    else:
        if standard_price != 0:
            price_deviation = abs(entered_price - standard_price) / standard_price
        if price_deviation > PRICE_TOLERANCE_PCT:
            price_flag = True
            direction  = "above" if entered_price > standard_price else "below"
            price_note = (f"FLAG – Entered price £{entered_price:.2f} is "
                          f"{price_deviation:.1%} {direction} standard £{standard_price:.2f}")

    # --- Check 2: Missing mandatory fields ---
    missing_fields = []
    if not cust_id or cust_id in ("nan", ""):
        missing_fields.append("Customer ID")
    if not po_ref or po_ref in ("nan", ""):
        missing_fields.append("PO Reference")
    if not uom or uom in ("nan", ""):
        missing_fields.append("Unit of Measure")
    if qty <= 0:
        missing_fields.append("Quantity")
    field_flag = len(missing_fields) > 0
    field_note = "OK – all mandatory fields present"
    if field_flag:
        field_note = f"FLAG – Missing: {', '.join(missing_fields)}"

    # --- Check 3: Line value arithmetic integrity ---
    if standard_price:
        expected_lv = round(qty * entered_price, 2)
        lv_variance = abs(line_value - expected_lv)
        lv_flag     = lv_variance > 0.01
        lv_note     = "OK" if not lv_flag else f"FLAG – Line value £{line_value:.2f} ≠ qty×price £{expected_lv:.2f}"
    else:
        lv_flag, lv_note = False, "N/A – no standard price"

    # --- Check 4: Date logic ---
    date_flag = False
    date_note = "OK"
    try:
        od = datetime.strptime(order_date, DATE_FORMAT)
        dd = datetime.strptime(delivery_date, DATE_FORMAT)
        if dd <= od:
            date_flag = True
            date_note = "FLAG – Delivery date is not after order date"
        elif (dd - od).days < 3:
            date_flag = True
            date_note = "WARN – Delivery lead time is less than 3 days"
    except ValueError:
        date_flag = True
        date_note = "WARN – Date format issue"

    # --- Overall status ---
    any_flag = price_flag or field_flag or lv_flag or date_flag
    if any_flag:
        errors_found += 1
        overall = "FAIL"
    else:
        overall = "PASS"

    variance = round(entered_price - (standard_price or entered_price), 2)

    results.append({
        "Sales Order No.":      order_no,
        "PO Reference":         po_ref,
        "Week":                 week,
        "Customer ID":          cust_id,
        "Customer Name":        cust_name,
        "Material Code":        mat_code,
        "Product Description":  description,
        "Qty":                  qty,
        "Entered Price (£)":    entered_price,
        "Standard Price (£)":   standard_price if standard_price else "N/A",
        "Price Variance (£)":   variance if standard_price else "N/A",
        "Deviation %":          price_deviation if standard_price else "N/A",
        "Price Check":          price_note,
        "Field Check":          field_note,
        "Line Value Check":     lv_note,
        "Date Check":           date_note,
        "Overall Result":       overall,
    })

pass_count = len(results) - errors_found
print(f"  ✓ Validation complete")
print(f"  ✓ PASS : {pass_count} orders ({pass_count/len(results)*100:.1f}%)")
print(f"  ✗ FAIL : {errors_found} orders ({errors_found/len(results)*100:.1f}%)")


# ── STEP 4: DISCREPANCY SUMMARY ───────────────────────────────────────────
print("\n[4/5] Generating discrepancy summary statistics...")

results_df = pd.DataFrame(results)

# Error type counts
price_errors = results_df["Price Check"].str.startswith("FLAG").sum()
field_errors = results_df["Field Check"].str.startswith("FLAG").sum()
lv_errors    = results_df["Line Value Check"].str.startswith("FLAG").sum()
date_errors  = results_df["Date Check"].str.startswith("FLAG|WARN").sum()

# Week-by-week error rate
weekly = results_df.groupby("Week").agg(
    Total=("Overall Result", "count"),
    Fails=("Overall Result", lambda x: (x == "FAIL").sum())
).reset_index()
weekly["Error Rate"] = weekly["Fails"] / weekly["Total"]

print(f"\n  Error breakdown:")
print(f"    Price deviations >5%     : {price_errors}")
print(f"    Missing mandatory fields  : {field_errors}")
print(f"    Line value discrepancies  : {lv_errors}")
print(f"    Date logic issues        : {date_errors}")

print(f"\n  Weekly error rate:")
for _, r in weekly.iterrows():
    bar = "█" * int(r["Error Rate"] * 40)
    print(f"    {r['Week']}  {r['Error Rate']:>6.1%}  {bar}")


# ── STEP 5: WRITE VALIDATION REPORT WORKBOOK ──────────────────────────────
print(f"\n[5/5] Writing validation report to '{OUTPUT_REPORT}'...")

wb_out = openpyxl.Workbook()

# ── Sheet A: Validation Results ──────────────────────────────────────────
ws_res = wb_out.active
ws_res.title = "Validation Results"
ws_res.sheet_view.showGridLines = False
ws_res.freeze_panes = "A3"

col_keys = list(results[0].keys())
col_widths = [16,14,8,12,22,13,26,8,16,16,15,12,38,28,28,28,16]

# Title
ws_res.merge_cells(f'A1:{get_column_letter(len(col_keys))}1')
t = ws_res["A1"]
t.value = f"ORDER VALIDATION REPORT  |  Run: {datetime.now().strftime('%d %B %Y %H:%M')}  |  Tolerance: {PRICE_TOLERANCE_PCT:.0%}"
t.font = hfont(size=12)
t.fill = pfill(NAVY)
t.alignment = centre()
ws_res.row_dimensions[1].height = 28

# Headers
for col_idx, (key, w) in enumerate(zip(col_keys, col_widths), 1):
    ws_res.column_dimensions[get_column_letter(col_idx)].width = w
    cell = ws_res.cell(row=2, column=col_idx, value=key)
    cell.font = hfont(size=9)
    cell.fill = pfill(BLUE)
    cell.alignment = centre()
    cell.border = thin(WHITE)
ws_res.row_dimensions[2].height = 28

# Data
for row_idx, rec in enumerate(results, 3):
    alt = row_idx % 2 == 0
    fail = rec["Overall Result"] == "FAIL"
    for col_idx, key in enumerate(col_keys, 1):
        val  = rec[key]
        cell = ws_res.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin()
        cell.font   = bfont(size=9)
        # Row background
        if fail:
            cell.fill = pfill(RED_FILL)
        else:
            cell.fill = pfill(ALT_ROW) if alt else pfill(WHITE)
        # Alignment
        cell.alignment = centre() if col_idx in {1,3,4,6,8,17} else left_align()
        # Number formats
        if key in {"Entered Price (£)", "Standard Price (£)", "Price Variance (£)"}:
            if isinstance(val, float):
                cell.number_format = "£#,##0.00"
                cell.alignment = right_al()
        if key == "Deviation %":
            if isinstance(val, float):
                cell.number_format = "0.0%"
                cell.alignment = centre()
        # Overall result colour
        if key == "Overall Result":
            if val == "PASS":
                cell.font = Font(name="Arial", size=9, bold=True, color=GREEN)
            else:
                cell.font = Font(name="Arial", size=9, bold=True, color=RED)
        # Check field colour coding
        if key in {"Price Check", "Field Check", "Line Value Check", "Date Check"}:
            if str(val).startswith("FLAG"):
                cell.font = Font(name="Arial", size=9, bold=True, color=RED)
            elif str(val).startswith("WARN"):
                cell.font = Font(name="Arial", size=9, bold=True, color=AMBER)
    ws_res.row_dimensions[row_idx].height = 16

# ── Sheet B: Executive Summary ───────────────────────────────────────────
ws_exec = wb_out.create_sheet("Executive Summary")
ws_exec.sheet_view.showGridLines = False
ws_exec.column_dimensions["A"].width = 4
ws_exec.column_dimensions["B"].width = 36
ws_exec.column_dimensions["C"].width = 20
ws_exec.column_dimensions["D"].width = 4

ws_exec.merge_cells("A1:D1")
t = ws_exec["A1"]
t.value = "VALIDATION EXECUTIVE SUMMARY"
t.font  = hfont(size=14)
t.fill  = pfill(NAVY)
t.alignment = centre()
ws_exec.row_dimensions[1].height = 32

summary_items = [
    ("Script Run Date",          datetime.now().strftime("%d %B %Y  %H:%M")),
    ("Total Orders Validated",   str(len(results))),
    ("Orders Passed",            f"{pass_count}  ({pass_count/len(results)*100:.1f}%)"),
    ("Orders Failed",            f"{errors_found}  ({errors_found/len(results)*100:.1f}%)"),
    ("Price Deviations (>5%)",   str(price_errors)),
    ("Missing Field Errors",     str(field_errors)),
    ("Line Value Discrepancies", str(lv_errors)),
    ("Date Logic Issues",        str(date_errors)),
    ("Price Tolerance Applied",  f"{PRICE_TOLERANCE_PCT:.0%} deviation threshold"),
    ("VK13 Records Checked",     str(len(pricing_lookup))),
    ("Validation Checks per Order", "4 independent checks"),
    ("Recommended Action",       "Review all FAIL rows; apply VA02 corrections; update VK13 conditions"),
]

for idx, (label, value) in enumerate(summary_items, 3):
    ws_exec.row_dimensions[idx].height = 22
    alt = idx % 2 == 0
    bg  = pfill(ALT_ROW) if alt else pfill(WHITE)
    lc  = ws_exec.cell(row=idx, column=2, value=label)
    lc.font = bfont(size=11, bold=True)
    lc.fill = bg
    lc.alignment = left_align()
    lc.border = thin()
    vc  = ws_exec.cell(row=idx, column=3, value=value)
    vc.font = bfont(size=11)
    vc.fill = bg
    vc.alignment = left_align()
    vc.border = thin()

# ── Sheet C: Weekly Error Trend ──────────────────────────────────────────
ws_trend = wb_out.create_sheet("Weekly Error Trend")
ws_trend.sheet_view.showGridLines = False
ws_trend.column_dimensions["A"].width = 12
ws_trend.column_dimensions["B"].width = 18
ws_trend.column_dimensions["C"].width = 14
ws_trend.column_dimensions["D"].width = 14

ws_trend.merge_cells("A1:D1")
t = ws_trend["A1"]
t.value = "WEEKLY ERROR RATE ANALYSIS"
t.font  = hfont(size=12)
t.fill  = pfill(NAVY)
t.alignment = centre()
ws_trend.row_dimensions[1].height = 28

trend_headers = ["Week", "Total Orders", "Failed Orders", "Error Rate"]
for col_idx, h in enumerate(trend_headers, 1):
    cell = ws_trend.cell(row=2, column=col_idx, value=h)
    cell.font = hfont(size=10)
    cell.fill = pfill(BLUE)
    cell.alignment = centre()
    cell.border = thin(WHITE)
ws_trend.row_dimensions[2].height = 24

for row_idx, (_, r) in enumerate(weekly.iterrows(), 3):
    alt = row_idx % 2 == 0
    bg  = pfill(ALT_ROW) if alt else pfill(WHITE)
    vals = [r["Week"], int(r["Total"]), int(r["Fails"]), r["Error Rate"]]
    for col_idx, val in enumerate(vals, 1):
        cell = ws_trend.cell(row=row_idx, column=col_idx, value=val)
        cell.font = bfont(size=10)
        cell.fill = bg
        cell.alignment = centre()
        cell.border = thin()
        if col_idx == 4:
            cell.number_format = "0.0%"
            if val > 0.08:
                cell.font = Font(name="Arial", size=10, bold=True, color=RED)
            elif val > 0.03:
                cell.font = Font(name="Arial", size=10, bold=True, color=AMBER)
            else:
                cell.font = Font(name="Arial", size=10, bold=True, color=GREEN)
    ws_trend.row_dimensions[row_idx].height = 20

# Tab colours
ws_res.sheet_properties.tabColor   = BLUE
ws_exec.sheet_properties.tabColor  = NAVY
ws_trend.sheet_properties.tabColor = GREEN

wb_out.save(OUTPUT_REPORT)

print(f"  ✓ Report saved: {OUTPUT_REPORT}")
print("\n" + "=" * 65)
print("VALIDATION COMPLETE")
print(f"  Orders checked : {len(results)}")
print(f"  Passed         : {pass_count}")
print(f"  Failed         : {errors_found}")
print(f"  Error rate     : {errors_found/len(results)*100:.1f}%")
print("=" * 65)
print("\nNext steps:")
print("  1. Open the Validation Report workbook")
print("  2. Review all FAIL rows in 'Validation Results' sheet")
print("  3. Apply corrections via VA02 (SAP change order)")
print("  4. Re-run script after corrections to confirm zero failures")
print("  5. Archive final report as audit evidence\n")
